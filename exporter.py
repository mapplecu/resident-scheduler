import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

MONTHS_COLS = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun"]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _stress_color(value: int, max_value: int) -> str:
    """
    Return an ARGB hex string (openpyxl format) interpolating green→yellow→red.
    value=0 → bright green; value=max_value → bright red.
    """
    if max_value == 0:
        ratio = 0.0
    else:
        ratio = min(1.0, value / max_value)

    if ratio < 0.5:
        # green → yellow
        t = ratio / 0.5
        r = int(255 * t)
        g = 200
        b = 0
    else:
        # yellow → red
        t = (ratio - 0.5) / 0.5
        r = 255
        g = int(200 * (1 - t))
        b = 0

    return f"FF{r:02X}{g:02X}{b:02X}"


def _compute_stress_rows(row: pd.Series, rotation_stress: dict) -> tuple:
    """
    Given one row of the master schedule, compute:
      - monthly_stress list  (int per month; 0 for electives)
      - cumulative_stress list (reset to 0 after each elective month)
      - total_annual_stress   (simple sum of monthly values)
      - peak_cumulative       (max of cumulative_stress list)
    """
    monthly = []
    for col in MONTHS_COLS:
        rot_name = str(row.get(col, ''))
        if 'elective' in rot_name.lower():
            monthly.append(0)
        else:
            monthly.append(rotation_stress.get(rot_name, 5))

    cumulative = []
    running = 0
    for m_idx, ms in enumerate(monthly):
        rot_name = str(row.get(MONTHS_COLS[m_idx], ''))
        if 'elective' in rot_name.lower():
            running = 0          # reset on elective
        else:
            running += ms
        cumulative.append(running)

    total = sum(monthly)
    peak  = max(cumulative) if cumulative else 0
    return monthly, cumulative, total, peak


# ── Main export function ──────────────────────────────────────────────────────

def generate_excel_bytes(schedule_df: pd.DataFrame, rotation_stress: dict = None) -> bytes:
    """
    Takes the master schedule DataFrame and an optional dict {rotation_name: stress_score}.
    Writes a multi-sheet Excel file and returns raw bytes.

    Sheet order:
      1. Master Schedule  – one row per resident, months as columns
      2. Rotations        – non-elective rotations × months (comma-sep residents)
      3+. Individual resident sheets with stress heatmap rows
    """
    if rotation_stress is None:
        rotation_stress = {}

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:

        # ── 1. Master Sheet ──────────────────────────────────────────────────
        if 'Year' in schedule_df.columns:
            schedule_df['_pgy_num'] = schedule_df['Year'].apply(
                lambda x: int(x.split('-')[1]) if '-' in x else 0
            )
            master_sorted = schedule_df.sort_values(by='_pgy_num', ascending=False).drop(columns=['_pgy_num'])
        else:
            master_sorted = schedule_df

        master_sorted.to_excel(writer, sheet_name='Master Schedule', index=False)

        # ── 2. Rotations Sheet ───────────────────────────────────────────────
        all_rotation_names = set()
        for col in MONTHS_COLS:
            if col in master_sorted.columns:
                for val in master_sorted[col].dropna():
                    if 'elective' not in str(val).lower():
                        all_rotation_names.add(str(val))

        sorted_rotations = sorted(all_rotation_names)

        rotation_rows = []
        for rot_name in sorted_rotations:
            row = {'Rotation': rot_name}
            for col in MONTHS_COLS:
                if col in master_sorted.columns:
                    mask = master_sorted[col].astype(str) == rot_name
                    assigned = master_sorted.loc[mask, 'Resident'].tolist()
                    row[col] = ', '.join(assigned) if assigned else ''
                else:
                    row[col] = ''
            rotation_rows.append(row)

        rotations_df = pd.DataFrame(rotation_rows, columns=['Rotation'] + MONTHS_COLS)
        rotations_df.to_excel(writer, sheet_name='Rotations', index=False)

        # ── 3. Individual Resident Sheets ────────────────────────────────────
        wb = writer.book

        for _, row in master_sorted.iterrows():
            res_name = row['Resident']

            # Compute stress metrics for this resident
            monthly_stress, cumulative_stress, total_annual, peak_cum = \
                _compute_stress_rows(row, rotation_stress)

            # Build the base month data rows
            resident_data = []
            for col in MONTHS_COLS:
                if col in master_sorted.columns:
                    resident_data.append({"Month": col, "Assigned Rotation": row[col]})

            res_df = pd.DataFrame(resident_data)

            safe_name = str(res_name)[:31]
            for ch in ['*', ':', '?', '/', '\\', '[', ']']:
                safe_name = safe_name.replace(ch, '')

            res_df.to_excel(writer, sheet_name=safe_name, index=False)

            ws = wb[safe_name]
            num_rows = len(resident_data)   # 12 month rows
            header_row = 1
            data_start = 2                  # row 2 = Jul

            # Append Stress rows below the month table
            stress_label_row   = header_row + num_rows + 2   # blank gap row
            cum_label_row      = stress_label_row + 1
            summary_row        = cum_label_row + 2

            # --- Monthly Stress row ---
            ws.cell(row=stress_label_row, column=1, value="Monthly Stress Score")
            ws.cell(row=stress_label_row, column=1).font = Font(bold=True)
            max_monthly = max(monthly_stress) if monthly_stress else 1

            for i, val in enumerate(monthly_stress):
                c = ws.cell(row=stress_label_row, column=i + 2, value=val)
                fill_color = _stress_color(val, max_monthly if max_monthly else 1)
                c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                c.font = Font(bold=True, color="FF000000")
                c.alignment = Alignment(horizontal="center")

            # --- Cumulative Stress row ---
            ws.cell(row=cum_label_row, column=1, value="Cumulative Stress")
            ws.cell(row=cum_label_row, column=1).font = Font(bold=True)
            max_cum = max(cumulative_stress) if cumulative_stress else 1

            for i, val in enumerate(cumulative_stress):
                c = ws.cell(row=cum_label_row, column=i + 2, value=val)
                fill_color = _stress_color(val, max_cum if max_cum else 1)
                c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                c.font = Font(bold=True, color="FF000000")
                c.alignment = Alignment(horizontal="center")

            # --- Summary cells ---
            ws.cell(row=summary_row, column=1, value="Total Annual Stress Score:")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=total_annual)
            ws.cell(row=summary_row, column=2).font = Font(bold=True)

            ws.cell(row=summary_row + 1, column=1, value="Peak Cumulative Stress:")
            ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value=peak_cum)
            ws.cell(row=summary_row + 1, column=2).font = Font(bold=True)

    return buffer.getvalue()
